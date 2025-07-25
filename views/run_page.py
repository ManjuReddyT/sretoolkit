import importlib
import json
import os
import subprocess
import uuid
from datetime import datetime

import streamlit as st
from werkzeug.utils import secure_filename
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Constants
UPLOAD_DIR = "uploads"
REPORTS_DIR = "reports"
CONFIG_FILE = "config.json"
SCRIPTS_DIR = "scripts"

# Ensure the directories exist
def ensure_directories():
    """Ensures that necessary directories for uploads, reports, and scripts exist."""
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(REPORTS_DIR, exist_ok=True)
    os.makedirs(SCRIPTS_DIR, exist_ok=True)

# Load the configuration
def load_config():
    """Loads the configuration from config.json."""
    try:
        with open(CONFIG_FILE, "r") as f:
            return json.load(f)
    except FileNotFoundError:
        st.error(f"{CONFIG_FILE} file not found! Please create it.")
        st.stop() # Stop the app if config is missing
    except json.JSONDecodeError:
        st.error(f"Error decoding {CONFIG_FILE}. Please check the file format.")
        st.stop() # Stop the app if config is malformed

# Render input fields for a selected script
def render_inputs(inputs):
    """
    Renders input fields based on the script's configuration.

    Args:
        inputs (list): A list of dictionaries defining the input fields.

    Returns:
        dict: A dictionary containing user-provided input values.
    """
    user_inputs = {}
    for input_field in inputs:
        if input_field["type"] == "file":
            # Use a unique key for each file uploader instance based on input_field name
            file_key = f"file_uploader_{input_field['name']}"
            user_inputs[input_field["name"]] = st.file_uploader(
                input_field["label"],
                key=file_key, # Use specific key
                accept_multiple_files=input_field.get("multiple", False) # Support multiple files
            )
        elif input_field["type"] == "text":
            user_inputs[input_field["name"]] = st.text_input(input_field["label"])
        elif input_field["type"] == "date":
            date_value = st.date_input(input_field["label"])
            time_value = st.time_input(f"Time - {input_field['label']}", key=f"time_{input_field['name']}")
            user_inputs[input_field["name"]] = {"date": date_value, "time": time_value}
    return user_inputs

# Format AI analysis sheet in an Excel workbook
def format_ai_sheet(workbook, sheet_name, title, content, config):
    """
    Formats and adds an AI analysis sheet to an Excel workbook.

    Args:
        workbook (openpyxl.workbook.workbook.Workbook): The Excel workbook object.
        sheet_name (str): The name for the AI analysis sheet.
        title (str): The title for the AI analysis report.
        content (str): The AI generated analysis content.
        config (dict): The script's configuration, containing excel formatting details.

    Returns:
        openpyxl.workbook.workbook.Workbook: The modified Excel workbook.
    """
    # Remove existing sheet if it exists
    if sheet_name in workbook.sheetnames:
        ai_sheet = workbook[sheet_name]
        workbook.remove(ai_sheet)

    ai_sheet = workbook.create_sheet(sheet_name)

    # Apply styling from config or use defaults
    style_config = config.get("excel_formatting", {})
    header_color = style_config.get("header_color", "1072BA")
    font_color = style_config.get("header_font_color", "FFFFFF")
    col_width = style_config.get("analysis_column_width", 120)

    # Set column widths
    ai_sheet.column_dimensions['A'].width = col_width # Assuming content is primarily in column A

    # Write header
    header = ai_sheet.cell(row=1, column=1, value=title)
    header.font = Font(bold=True, size=14, color=font_color)
    header.fill = PatternFill("solid", fgColor=header_color)
    header.alignment = Alignment(horizontal="center", vertical="center")
    # For a single content column, merging A1 is sufficient, adjust if header spans multiple data columns
    # ai_sheet.merge_cells('A1:C1') # Removed this merge as content is primarily in col A

    # Write content with formatting
    current_row = 3
    # Split by double newline for paragraphs, then strip to remove empty ones
    for paragraph in content.split('\n\n'):
        paragraph_stripped = paragraph.strip()
        if paragraph_stripped:
            cell = ai_sheet.cell(row=current_row, column=1, value=paragraph_stripped)

            # Apply special formatting for certain content
            if any(word in paragraph_stripped.lower() for word in ['recommendation', 'suggestion', 'action item']):
                cell.fill = PatternFill("solid", fgColor=style_config.get("highlight_colors", {}).get("recommendation", "E6FFE6"))
            elif any(word in paragraph_stripped.lower() for word in ['warning', 'critical', 'error', 'issue']):
                cell.fill = PatternFill("solid", fgColor=style_config.get("highlight_colors", {}).get("warning", "FFE6E6"))

            cell.alignment = Alignment(wrap_text=True, vertical="top")
            current_row += 1

    return workbook

# Build the AI prompt safely
def build_ai_prompt(prompt_template, content):
    """
    Constructs the AI prompt based on a template and content.

    Args:
        prompt_template (dict): A dictionary defining the prompt structure.
        content (str): The data content to be included in the prompt.

    Returns:
        str: The formatted AI prompt.
    """
    system_role = prompt_template.get("system_role", "You are an expert technical analyst.")
    task_desc = prompt_template.get("task_description", "Please analyze this data:")
    analysis_points = prompt_template.get("analysis_points", ['key insights'])
    output_format = prompt_template.get("output_format", "Provide your analysis in clear, organized format.")

    analysis_points_str = '\n'.join(f'- {point}' for point in analysis_points)

    prompt = (
        f"{system_role}\n\n"
        f"{task_desc}\n\n"
        "Here is the relevant data:\n"
        "---\n"
        f"{content}\n"
        "---\n\n"
        "Please analyze this data focusing on:\n"
        f"{analysis_points_str}\n\n"
        f"{output_format}"
    )
    return prompt

# Run script with optional AI analysis
def run_script(script_data, user_inputs, ollama_api_url, ollama_model_name, global_ai_analysis):
    """
    Executes the selected script and performs AI analysis if configured.

    Args:
        script_data (dict): Dictionary containing details of the selected script.
        user_inputs (dict): Dictionary of user-provided inputs for the script.
        ollama_api_url (str): The URL for the Ollama API.
        ollama_model_name (str): The name of the Ollama model to use.
        global_ai_analysis (bool): Flag indicating if global AI analysis is enabled.

    Returns:
        str or None: The path to the generated output file if successful, otherwise None.
    """
    unique_id = str(uuid.uuid4())[:8]
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

    input_data = {}
    perform_ai_analysis_by_run_page = False # Flag to indicate if run_page.py should do AI analysis

    # Prepare input data and save uploaded files
    for field in script_data["inputs"]:
        if field["type"] == "file":
            uploaded_files_input = user_inputs.get(field["name"]) # This could be a single file or a list of files
            
            if uploaded_files_input:
                saved_file_paths_for_script = []
                
                # Handle single vs. multiple file uploads for saving
                if isinstance(uploaded_files_input, list):
                    for i, uploaded_file_obj in enumerate(uploaded_files_input):
                        filename = f"{secure_filename(uploaded_file_obj.name).split('.')[0]}_{unique_id}_{timestamp}_{i}{os.path.splitext(uploaded_file_obj.name)[1]}"
                        file_path = os.path.join(UPLOAD_DIR, filename)
                        with open(file_path, "wb") as f:
                            f.write(uploaded_file_obj.getbuffer())
                        saved_file_paths_for_script.append(file_path)
                    input_data[field["name"]] = saved_file_paths_for_script
                else:
                    uploaded_file_obj = uploaded_files_input
                    filename = f"{secure_filename(uploaded_file_obj.name).split('.')[0]}_{unique_id}_{timestamp}{os.path.splitext(uploaded_file_obj.name)[1]}"
                    file_path = os.path.join(UPLOAD_DIR, filename)
                    with open(file_path, "wb") as f:
                        f.write(uploaded_file_obj.getbuffer())
                    input_data[field["name"]] = file_path

                # Determine if AI analysis should be performed by run_page.py
                # This flag is set if global AI is on AND this specific input field requests AI analysis
                if global_ai_analysis and field.get("ai-analysis"):
                    perform_ai_analysis_by_run_page = True
        else:
            value = user_inputs.get(field["name"])
            if isinstance(value, dict) and "date" in value and "time" in value:
                if value["date"] and value["time"]:
                    input_data[field["name"]] = f"{value['date']}T{value['time']}:00Z"
                else:
                    input_data[field["name"]] = None
            else:
                input_data[field["name"]] = value

    # Define the output file path for subprocess-generated reports
    output_file = os.path.join(REPORTS_DIR, f"{script_data['name']}_{unique_id}_{timestamp}.{script_data['output_type']}")

    with st.spinner("Running the script..."):
        try:
            result = subprocess.run(
                ["python3", script_data["script_path"], json.dumps(input_data), output_file],
                capture_output=True, text=True, check=True
            )
            st.success("Script executed successfully!")

            # AI Analysis Section for run_page.py to perform analysis
            if perform_ai_analysis_by_run_page:
                # The AI content source is always derived from the output_file as _summary.txt
                ai_content_source_path = output_file.replace(f".{script_data['output_type']}", '_summary.txt')

                if os.path.exists(ai_content_source_path):
                    st.subheader("AI Analysis")
                    with st.spinner("Running AI analysis..."):
                        full_ai_response = ""
                        try:
                            # Read a snippet of the file designated for AI prompt
                            log_content_for_ai = ""
                            with open(ai_content_source_path, 'r', encoding='utf-8', errors='ignore') as f:
                                log_content_for_ai = f.read(4096) # Read first 4KB for context
                                if not log_content_for_ai.strip():
                                    st.warning(f"AI source file '{os.path.basename(ai_content_source_path)}' is empty. AI analysis may not be meaningful.")

                            prompt_template = script_data.get("ai_prompt", {})
                            prompt = build_ai_prompt(prompt_template, log_content_for_ai)

                            headers = {'Content-Type': 'application/json'}
                            data = {
                                "model": ollama_model_name,
                                "prompt": prompt,
                                "stream": True,
                                "options": {"num_predict": 2048}
                            }

                            response = requests.post(ollama_api_url, headers=headers, json=data, stream=True)
                            response.raise_for_status()

                            ai_response_placeholder = st.empty()
                            for chunk in response.iter_content(chunk_size=None):
                                if chunk:
                                    try:
                                        json_data = json.loads(chunk.decode('utf-8'))
                                        full_ai_response += json_data.get("response", "")
                                        ai_response_placeholder.markdown(full_ai_response)
                                    except json.JSONDecodeError:
                                        pass

                            st.success("AI analysis complete!")

                            # If the main output is XLSX, add AI analysis to it
                            if script_data['output_type'] == 'xlsx':
                                try:
                                    workbook = openpyxl.load_workbook(output_file)
                                    workbook = format_ai_sheet(
                                        workbook,
                                        "AI Analysis",
                                        f"AI Analysis Report - {script_data['display_name']}",
                                        full_ai_response,
                                        script_data
                                    )
                                    workbook.save(output_file)
                                    st.success("Formatted AI analysis added to the report!")
                                except Exception as e:
                                    st.error(f"Error formatting Excel file with AI analysis: {e}")

                        except requests.exceptions.ConnectionError:
                            st.error("AI analysis failed: Could not connect to Ollama API. Ensure Ollama is running and accessible.")
                        except requests.exceptions.RequestException as e:
                            st.error(f"AI analysis failed due to a request error: {e}")
                        except Exception as e:
                            st.error(f"An unexpected error occurred during AI analysis: {e}")
                else:
                    st.warning(f"AI analysis was enabled but the source file for AI content was not found at '{ai_content_source_path}'. Ensure the script generates the expected '_summary.txt' file.")

            return output_file
        except subprocess.CalledProcessError as e:
            st.error(f"Script execution error: {e.stderr}")
            if e.stdout:
                st.code(e.stdout, language="python")
            return None

# Validate inputs
def validate_inputs(script_data, user_inputs):
    """
    Validates user inputs against the script's defined input requirements.

    Args:
        script_data (dict): Dictionary containing details of the selected script.
        user_inputs (dict): Dictionary of user-provided input values.

    Returns:
        bool: True if all inputs are valid, False otherwise.
    """
    for field in script_data["inputs"]:
        field_name = field["name"]
        value = user_inputs.get(field_name)

        if field.get("required"): # Only check if required is true
            if field["type"] == "file" and field.get("multiple"):
                if not value: # For multiple files, value is a list (can be empty)
                    st.error(f"{field['label']} is required.")
                    return False
            elif not value: # For single files or other types, check if value is None/empty
                st.error(f"{field['label']} is required.")
                return False

        # File validation (for single or multiple files)
        if field["type"] == "file" and value:
            files_to_check = value if isinstance(value, list) else [value]
            for uploaded_file in files_to_check:
                extension = os.path.splitext(uploaded_file.name)[1].lower()
                if "allowed_extensions" in field and extension not in field["allowed_extensions"]:
                    st.error(f"{field['label']} must be one of the following: {', '.join(field['allowed_extensions'])}")
                    return False

        if field["type"] == "text" and value:
            if "max_length" in field and len(value) > field["max_length"]:
                st.error(f"{field['label']} cannot exceed {field['max_length']} characters.")
                return False
            if "regex" in field:
                import re
                if not re.match(field["regex"], value):
                    st.error(f"{field['label']} must match the format: {field['regex']}")
                    return False

        if field["type"] == "date" and value and value["date"] is not None: # Ensure date_value is not None
            date_value = value["date"]
            if "must_be_future" in field and field["must_be_future"] and date_value <= datetime.now().date():
                st.error(f"{field['label']} must be a future date.")
                return False

    return True

# Main page function
def run_page():
    """Main function to run the Streamlit application page."""
    ensure_directories()
    config = load_config()

    st.title("SRE Toolkit")

    # Get default Ollama settings from config or use hardcoded defaults
    ollama_defaults = config.get("ollama_defaults", {})
    default_ollama_api_url = ollama_defaults.get("api_url", "http://localhost:11434/api/generate")
    default_ollama_model_name = ollama_defaults.get("model_name", "deepseek-r1:8b")

    # UI inputs for Ollama API and Model moved to sidebar
    with st.sidebar:
        st.header("AI Model Settings")
        ollama_api_url_input = st.text_input(
            "Ollama API URL",
            value=default_ollama_api_url,
            help="Enter the URL for your Ollama API endpoint (e.g., http://localhost:11434/api/generate)"
        )
        ollama_model_name_input = st.text_input(
            "Ollama Model Name",
            value=default_ollama_model_name,
            help="Enter the name of the Ollama model to use (e.g., deepseek-r1:8b)"
        )
        global_ai_analysis = st.checkbox("Enable AI Analysis", value=True, help="Globally enable or disable AI analysis for scripts that support it.")

    st.divider()

    script_names = [script["display_name"] for script in config["scripts"]]
    script_selection = st.selectbox("Select a script", script_names)

    if "previous_script" not in st.session_state:
        st.session_state.previous_script = None

    selected_script = next(script for script in config["scripts"] if script["display_name"] == script_selection)

    # Store script-specific and global configuration in session state for scripts to access
    # This is crucial for "streamlit_app" types to access settings like "aianalysis"
    st.session_state["config"] = {
        "ollama_api_url": ollama_api_url_input,
        "ollama_model_name": ollama_model_name_input,
        "aianalysis": global_ai_analysis and any( # Check if global AI is on AND any input field specifies ai-analysis
            input_field.get("ai-analysis", False)
            for input_field in selected_script.get("inputs", [])
        )
    }

    with st.expander(f"ℹ️ Help: {selected_script.get('name', 'Script')}"):
        st.markdown(selected_script.get("help_text", "No help information available."))
        if "ai_prompt" in selected_script:
            st.markdown("**AI Analysis Will Provide:**")
            st.markdown("".join(f"- {point}\n" for point in selected_script["ai_prompt"].get("analysis_points", [])))

    st.divider()
    st.session_state.script_name = selected_script["name"]

    # Reset file uploader state when script selection changes
    if st.session_state.previous_script != st.session_state.script_name:
        st.session_state.previous_script = st.session_state.script_name
        # Iterate through all configured scripts to find all possible file uploader keys
        for script_cfg in config["scripts"]:
            for input_field in script_cfg.get("inputs", []):
                if input_field["type"] == "file":
                    file_key = f"file_uploader_{input_field['name']}"
                    if file_key in st.session_state:
                        del st.session_state[file_key]

    st.header(f"SRE Toolkit - {selected_script['display_name']}")
    st.write(selected_script['description'])

    # Conditional rendering based on output_type
    if selected_script["output_type"] == "streamlit_app": # Use "streamlit_app" as per config.json
        try:
            # Dynamically import and run the Streamlit app script's main function
            spec = importlib.util.spec_from_file_location("dynamic_module", selected_script["script_path"])
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            module.main() # Call the main function of the loaded Streamlit app
        except Exception as e:
            st.error(f"Error loading or running Streamlit app '{selected_script['display_name']}': {e}")
            st.exception(e) # Show full traceback for debugging
    else:
        # Existing logic for scripts that generate files (xlsx, csv, html, main)
        user_inputs = render_inputs(selected_script["inputs"])

        if st.button("Run Analysis"):
            if validate_inputs(selected_script, user_inputs):
                output_file = run_script(
                    selected_script, user_inputs,
                    ollama_api_url_input, ollama_model_name_input,
                    global_ai_analysis # Pass global AI flag
                )
                if output_file:
                    with open(output_file, "rb") as f:
                        file_data = f.read()

                    # Determine MIME type dynamically for download button
                    mime_type_map = {
                        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "csv": "text/csv",
                        "html": "text/html",
                        "log": "text/plain",
                        "txt": "text/plain",
                        "json": "application/json",
                        "har": "application/json", # HAR files are JSON
                        "main": "application/octet-stream" # Fallback for 'main' if it produces a file
                    }
                    download_mime_type = mime_type_map.get(selected_script['output_type'], "application/octet-stream")

                    st.download_button(
                        "Download Full Report",
                        file_data,
                        file_name=os.path.basename(output_file),
                        mime=download_mime_type
                    )

# Run the main page function
run_page()